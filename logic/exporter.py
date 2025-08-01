import os
import pandas as pd
import openpyxl
from datetime import datetime

# Generador de archivo exportador de control interno e importacion masiva
def exportar_control_interno(df_total, df_total_per_liq, df_balance, df_summary,df_tesoreria, df_masive_import, path_salida,callback_progress):
    
    # Crear directorio de salida si no existe
    os.makedirs("salida", exist_ok=True)

    # Intentar borrar el archivo si ya existe
    if os.path.exists(path_salida):
        try:
            os.remove(path_salida)
        except PermissionError:
            raise PermissionError(f"No se puede sobrescribir el archivo: {path_salida}. Cerralo si está abierto.")

    callback_progress("✅ Exportando archivo... (98%)")
    
    # Crear archivo Excel de importación masiva
    df_masive_import.to_excel("salida/importacion_masiva.xlsx", index=False)

    # Limpieza general de df_balance
    df_balance_export = df_balance.reset_index()
    df_balance_export.columns = [str(c).strip() for c in df_balance_export.columns]
    df_balance_export = df_balance_export.applymap(
        lambda x: x if isinstance(x, (int, float)) or pd.isna(x) else str(x)
    )

    callback_progress("✅ Exportando archivo... (99%)")

    # Crear archivo Excel de control interno
    with pd.ExcelWriter(path_salida, engine='openpyxl', datetime_format='DD/MM/YYYY') as writer:
        # Hoja total   
        df_total.to_excel(writer, index=False, sheet_name="total")
        # Hoja total agrupado por liquidacion
        df_total_per_liq.to_excel(writer, index=False, sheet_name="total agrupado por liquidacion")
        # Hoja balance
        df_balance_export.to_excel(writer, index=False, sheet_name="TD LIQ")
        # Hoja resumen
        df_summary.to_excel(writer, index=True, sheet_name="Resumen",merge_cells=True)
        # Hoja tesoreria
        df_tesoreria.to_excel(writer, index=False, sheet_name="Tesoreria")

        callback_progress("✅ Exportando archivo... (100%)")

        # --- FORMATO FECHA: total ---
        ws_total = writer.sheets["total"]
        for col in ['Fecha de Liquidación', 'Período de Liquidación']:
            if col in df_total.columns:
                col_idx = df_total.columns.get_loc(col) + 1
                for row in range(2, len(df_total) + 2):
                    ws_total.cell(row=row, column=col_idx).number_format = 'DD/MM/YYYY'

        # --- FORMATO FECHA: agrupado ---
        ws_agrupado = writer.sheets["total agrupado por liquidacion"]
        for col in ['Fecha de Factura', 'Vencimiento de Pago', 'Fecha de Emisión']:
            if col in df_total_per_liq.columns:
                col_idx = df_total_per_liq.columns.get_loc(col) + 1
                for row in range(2, len(df_total_per_liq) + 2):
                    ws_agrupado.cell(row=row, column=col_idx).number_format = 'DD/MM/YYYY'
        
    return path_salida, df_total
